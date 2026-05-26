from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query, UploadFile, File
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from collections import Counter, defaultdict
import os, uuid, logging, bcrypt, jwt, random

# ---------- DB ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = "HS256"
SUPER_ADMIN_KEY = os.environ.get('SUPER_ADMIN_KEY', 'VIQSO-MASTER-2026-XKL9PQR4')
DEFAULT_ORG_ID = "default-org-001"
DEFAULT_ACCESS_KEY = os.environ.get('DEFAULT_ACCESS_KEY', 'VIQSO-2026')
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Voter CRM API")
api = APIRouter(prefix="/api")


# ---------- Helpers ----------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def create_access_token(user_id: str, email: str, role: str, org_id: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role, "org_id": org_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def extract_surname(name: str) -> str:
    if not name:
        return ""
    parts = [p for p in name.strip().split() if p]
    return parts[-1].title() if parts else ""


def auto_family_id(address: str, surname: str) -> Optional[str]:
    if not address or not surname:
        return None
    import hashlib
    norm = (address.strip().lower() + "|" + surname.strip().lower()).encode()
    return "fam-" + hashlib.md5(norm).hexdigest()[:12]


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        h = request.headers.get("Authorization", "")
        if h.startswith("Bearer "):
            token = h[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "User not found")
    if not user.get("org_id"):
        user["org_id"] = payload.get("org_id") or DEFAULT_ORG_ID
    org = await db.organizations.find_one({"id": user["org_id"], "active": True})
    if not org:
        raise HTTPException(403, "Organization inactive or removed")
    user["org_name"] = org.get("name")
    return user


def org_filter(user: dict) -> Dict[str, Any]:
    return {"org_id": user["org_id"]}


def worker_booth_ids(user: dict) -> List[str]:
    booth_ids = list(user.get("assigned_booth_ids") or [])
    if user.get("booth_id"):
        booth_ids.append(user["booth_id"])
    return list(set(booth_ids))


def require_super_admin(request: Request):
    key = request.headers.get("X-Super-Admin-Key", "")
    if key != SUPER_ADMIN_KEY:
        raise HTTPException(403, "Invalid super admin key")
    return True


async def record_login_attempt(identifier: str, success: bool):
    if success:
        await db.login_attempts.delete_many({"identifier": identifier})
        return
    await db.login_attempts.insert_one({
        "identifier": identifier,
        "at": datetime.now(timezone.utc).isoformat(),
    })


async def is_locked_out(identifier: str) -> bool:
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
    count = await db.login_attempts.count_documents({
        "identifier": identifier,
        "at": {"$gte": cutoff},
    })
    return count >= MAX_LOGIN_ATTEMPTS


def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(403, f"Requires one of: {roles}")
        return user
    return checker


# ---------- Models ----------
class LoginInput(BaseModel):
    email: str
    password: str
    access_key: str

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str  # admin | supervisor | worker
    phone: Optional[str] = None
    booth_id: Optional[str] = None
    assigned_booth_ids: Optional[List[str]] = []

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    booth_id: Optional[str] = None
    assigned_booth_ids: Optional[List[str]] = None
    password: Optional[str] = None

class BoothCreate(BaseModel):
    name: str
    booth_number: str
    ward: str
    constituency: str
    location: Optional[str] = None
    target_voters: int = 0
    supervisor_id: Optional[str] = None
    assigned_workers: Optional[List[str]] = []

class BoothUpdate(BaseModel):
    name: Optional[str] = None
    booth_number: Optional[str] = None
    ward: Optional[str] = None
    constituency: Optional[str] = None
    location: Optional[str] = None
    target_voters: Optional[int] = None
    supervisor_id: Optional[str] = None
    assigned_workers: Optional[List[str]] = None

class VoterCreate(BaseModel):
    booth_id: str
    name: str
    voter_id_number: Optional[str] = None
    age: Optional[int] = None
    gender: Optional[str] = None  # male/female/other
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    caste: Optional[str] = None
    religion: Optional[str] = None
    occupation: Optional[str] = None
    political_preference: Optional[str] = None  # supporter/neutral/opposition/undecided
    sentiment: Optional[str] = None  # positive/neutral/negative
    issues: Optional[List[str]] = []
    likely_to_vote: Optional[bool] = None
    notes: Optional[str] = None
    custom_fields: Optional[Dict[str, Any]] = {}
    family_id: Optional[str] = None
    surname: Optional[str] = None

class VoterUpdate(BaseModel):
    name: Optional[str] = None
    voter_id_number: Optional[str] = None
    age: Optional[int] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    caste: Optional[str] = None
    religion: Optional[str] = None
    occupation: Optional[str] = None
    political_preference: Optional[str] = None
    sentiment: Optional[str] = None
    issues: Optional[List[str]] = None
    likely_to_vote: Optional[bool] = None
    notes: Optional[str] = None
    custom_fields: Optional[Dict[str, Any]] = None
    booth_id: Optional[str] = None
    family_id: Optional[str] = None
    surname: Optional[str] = None

class VisitCreate(BaseModel):
    booth_id: str
    worker_id: str
    scheduled_date: str
    notes: Optional[str] = None

class VisitUpdate(BaseModel):
    status: Optional[str] = None  # scheduled/completed/missed
    notes: Optional[str] = None
    voters_contacted: Optional[int] = None


# ---------- AUTH ----------
@api.post("/auth/login")
async def login(body: LoginInput, request: Request, response: Response):
    email = body.email.lower().strip()
    access_key = (body.access_key or "").strip()
    client_ip = (request.client.host if request.client else "unknown")
    identifier = f"{client_ip}:{email}"

    if await is_locked_out(identifier):
        raise HTTPException(429, f"Too many failed attempts. Try again in {LOCKOUT_MINUTES} minutes.")

    if not access_key:
        await record_login_attempt(identifier, False)
        raise HTTPException(401, "Access key is required")

    org = await db.organizations.find_one({"access_key": access_key, "active": True})
    if not org:
        await record_login_attempt(identifier, False)
        raise HTTPException(401, "Invalid access key")

    user = await db.users.find_one({"email": email, "org_id": org["id"]})
    if not user or not verify_password(body.password, user["password_hash"]):
        await record_login_attempt(identifier, False)
        raise HTTPException(401, "Invalid email or password for this organization")

    await record_login_attempt(identifier, True)
    token = create_access_token(user["id"], user["email"], user["role"], user["org_id"])
    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    user.pop("_id", None); user.pop("password_hash", None)
    user["org_name"] = org.get("name")
    return {"user": user, "access_token": token, "org": {"id": org["id"], "name": org["name"], "party_name": org.get("party_name")}}

@api.post("/auth/logout")
async def logout(response: Response, _user=Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


# ---------- USERS (Admin) ----------
@api.get("/users")
async def list_users(user=Depends(require_roles("admin", "supervisor"))):
    users = await db.users.find({"org_id": user["org_id"]}, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return users

@api.post("/users")
async def create_user(body: UserCreate, admin=Depends(require_roles("admin"))):
    email = body.email.lower().strip()
    if await db.users.find_one({"email": email, "org_id": admin["org_id"]}):
        raise HTTPException(400, "Email already exists in this organization")
    if body.role not in ["admin", "supervisor", "worker"]:
        raise HTTPException(400, "Invalid role")
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hash_password(body.password),
        "name": body.name,
        "role": body.role,
        "phone": body.phone,
        "booth_id": body.booth_id,
        "assigned_booth_ids": body.assigned_booth_ids or [],
        "active": True,
        "org_id": admin["org_id"],
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    doc.pop("_id", None); doc.pop("password_hash", None)
    return doc

@api.patch("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, admin=Depends(require_roles("admin"))):
    target = await db.users.find_one({"id": user_id, "org_id": admin["org_id"]})
    if not target:
        raise HTTPException(404, "User not found")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    if not updates:
        raise HTTPException(400, "No updates")
    await db.users.update_one({"id": user_id, "org_id": admin["org_id"]}, {"$set": updates})
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return user

@api.delete("/users/{user_id}")
async def delete_user(user_id: str, admin=Depends(require_roles("admin"))):
    if user_id == admin["id"]:
        raise HTTPException(400, "Cannot delete yourself")
    res = await db.users.delete_one({"id": user_id, "org_id": admin["org_id"]})
    return {"deleted": res.deleted_count}


# ---------- BOOTHS ----------
@api.get("/booths")
async def list_booths(user=Depends(get_current_user)):
    q = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        q["id"] = {"$in": worker_booth_ids(user)}
    booths = await db.booths.find(q, {"_id": 0}).sort("booth_number", 1).to_list(1000)
    for b in booths:
        b["voters_surveyed"] = await db.voters.count_documents({"booth_id": b["id"]})
    return booths

@api.post("/booths")
async def create_booth(body: BoothCreate, admin=Depends(require_roles("admin"))):
    doc = {"id": str(uuid.uuid4()), **body.model_dump(), "org_id": admin["org_id"], "created_at": now_iso()}
    await db.booths.insert_one(doc)
    doc.pop("_id", None)
    doc["voters_surveyed"] = 0
    return doc

@api.get("/booths/{booth_id}")
async def get_booth(booth_id: str, user=Depends(get_current_user)):
    b = await db.booths.find_one({"id": booth_id, "org_id": user["org_id"]}, {"_id": 0})
    if not b:
        raise HTTPException(404, "Booth not found")
    if user["role"] == "worker" and booth_id not in worker_booth_ids(user):
        raise HTTPException(403, "Not assigned to this booth")
    b["voters_surveyed"] = await db.voters.count_documents({"booth_id": booth_id})
    workers = await db.users.find({
        "org_id": user["org_id"],
        "$or": [{"booth_id": booth_id}, {"assigned_booth_ids": booth_id}]
    }, {"_id": 0, "password_hash": 0}).to_list(1000)
    b["workers"] = workers
    return b

@api.patch("/booths/{booth_id}")
async def update_booth(booth_id: str, body: BoothUpdate, user=Depends(require_roles("admin", "supervisor"))):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates")
    res = await db.booths.update_one({"id": booth_id, "org_id": user["org_id"]}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Booth not found")
    b = await db.booths.find_one({"id": booth_id}, {"_id": 0})
    return b

@api.delete("/booths/{booth_id}")
async def delete_booth(booth_id: str, admin=Depends(require_roles("admin"))):
    await db.booths.delete_one({"id": booth_id, "org_id": admin["org_id"]})
    return {"ok": True}


# ---------- VOTERS / SURVEYS ----------
@api.get("/voters")
async def list_voters(
    user=Depends(get_current_user),
    booth_id: Optional[str] = None,
    search: Optional[str] = None,
    political_preference: Optional[str] = None,
    sentiment: Optional[str] = None,
    limit: int = 200,
):
    q: Dict[str, Any] = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        q["booth_id"] = {"$in": worker_booth_ids(user)}
    if booth_id:
        if user["role"] == "worker" and booth_id not in worker_booth_ids(user):
            raise HTTPException(403, "Not assigned to this booth")
        q["booth_id"] = booth_id
    if political_preference:
        q["political_preference"] = political_preference
    if sentiment:
        q["sentiment"] = sentiment
    if search:
        q["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"voter_id_number": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
        ]
    voters = await db.voters.find(q, {"_id": 0}).sort("surveyed_at", -1).limit(limit).to_list(limit)
    return voters

@api.post("/voters")
async def create_voter(body: VoterCreate, user=Depends(get_current_user)):
    # Verify booth exists in user's org
    booth = await db.booths.find_one({"id": body.booth_id, "org_id": user["org_id"]})
    if not booth:
        raise HTTPException(404, "Booth not found in your organization")
    # Worker booth-scope enforcement
    if user["role"] == "worker" and body.booth_id not in worker_booth_ids(user):
        raise HTTPException(403, "You are not assigned to this booth")

    data = body.model_dump()
    surname = data.get("surname") or extract_surname(data.get("name", ""))
    family_id = data.get("family_id") or auto_family_id(data.get("address", ""), surname)
    doc = {
        "id": str(uuid.uuid4()),
        **data,
        "surname": surname,
        "family_id": family_id,
        "org_id": user["org_id"],
        "surveyed_by": user["id"],
        "surveyed_by_name": user["name"],
        "surveyed_at": now_iso(),
    }
    await db.voters.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/voters/{voter_id}")
async def get_voter(voter_id: str, user=Depends(get_current_user)):
    v = await db.voters.find_one({"id": voter_id, "org_id": user["org_id"]}, {"_id": 0})
    if not v:
        raise HTTPException(404, "Not found")
    if user["role"] == "worker" and v.get("booth_id") not in worker_booth_ids(user):
        raise HTTPException(403, "Not assigned to this booth")
    return v

@api.patch("/voters/{voter_id}")
async def update_voter(voter_id: str, body: VoterUpdate, user=Depends(get_current_user)):
    current = await db.voters.find_one({"id": voter_id, "org_id": user["org_id"]}, {"_id": 0})
    if not current:
        raise HTTPException(404, "Not found")
    if user["role"] == "worker":
        wb = worker_booth_ids(user)
        if current.get("booth_id") not in wb:
            raise HTTPException(403, "Not assigned to this booth")
        # If target booth changing, verify new booth is also in scope
        if body.booth_id and body.booth_id not in wb:
            raise HTTPException(403, "Target booth not in your scope")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if "name" in updates and not updates.get("surname"):
        updates["surname"] = extract_surname(updates["name"])
    if not updates.get("family_id"):
        addr = updates.get("address", current.get("address", ""))
        sn = updates.get("surname", current.get("surname", "")) or extract_surname(updates.get("name", current.get("name", "")))
        fid = auto_family_id(addr, sn)
        if fid:
            updates["family_id"] = fid
    updates["updated_at"] = now_iso()
    await db.voters.update_one({"id": voter_id, "org_id": user["org_id"]}, {"$set": updates})
    v = await db.voters.find_one({"id": voter_id}, {"_id": 0})
    return v

@api.delete("/voters/{voter_id}")
async def delete_voter(voter_id: str, user=Depends(require_roles("admin", "supervisor"))):
    await db.voters.delete_one({"id": voter_id, "org_id": user["org_id"]})
    return {"ok": True}


# ---------- VISITS ----------
@api.get("/visits")
async def list_visits(user=Depends(get_current_user), booth_id: Optional[str] = None):
    q: Dict[str, Any] = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        q["worker_id"] = user["id"]
    if booth_id:
        q["booth_id"] = booth_id
    visits = await db.visits.find(q, {"_id": 0}).sort("scheduled_date", 1).to_list(500)
    return visits

@api.post("/visits")
async def create_visit(body: VisitCreate, user=Depends(require_roles("admin", "supervisor"))):
    # Booth must belong to org
    if not await db.booths.find_one({"id": body.booth_id, "org_id": user["org_id"]}):
        raise HTTPException(404, "Booth not in your organization")
    # Worker must belong to org
    if not await db.users.find_one({"id": body.worker_id, "org_id": user["org_id"]}):
        raise HTTPException(404, "Worker not in your organization")
    doc = {
        "id": str(uuid.uuid4()),
        **body.model_dump(),
        "status": "scheduled",
        "voters_contacted": 0,
        "org_id": user["org_id"],
        "created_at": now_iso(),
    }
    await db.visits.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/visits/{visit_id}")
async def update_visit(visit_id: str, body: VisitUpdate, user=Depends(get_current_user)):
    visit = await db.visits.find_one({"id": visit_id, "org_id": user["org_id"]})
    if not visit:
        raise HTTPException(404, "Not found")
    if user["role"] == "worker" and visit.get("worker_id") != user["id"]:
        raise HTTPException(403, "Not your visit")
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates")
    await db.visits.update_one({"id": visit_id, "org_id": user["org_id"]}, {"$set": updates})
    v = await db.visits.find_one({"id": visit_id}, {"_id": 0})
    return v

@api.delete("/visits/{visit_id}")
async def delete_visit(visit_id: str, user=Depends(require_roles("admin", "supervisor"))):
    await db.visits.delete_one({"id": visit_id, "org_id": user["org_id"]})
    return {"ok": True}


# ---------- ANALYTICS ----------
@api.get("/analytics/overview")
async def analytics_overview(user=Depends(get_current_user)):
    voter_filter = {"org_id": user["org_id"]}
    booth_filter = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        wb = worker_booth_ids(user)
        voter_filter["booth_id"] = {"$in": wb}
        booth_filter["id"] = {"$in": wb}

    total_voters = await db.voters.count_documents(voter_filter)
    total_booths = await db.booths.count_documents(booth_filter)
    total_workers = await db.users.count_documents({"role": "worker", "org_id": user["org_id"]})

    booths = await db.booths.find(booth_filter, {"_id": 0, "target_voters": 1}).to_list(1000)
    total_target = sum(b.get("target_voters", 0) for b in booths)

    pref_cursor = db.voters.aggregate([
        {"$match": voter_filter},
        {"$group": {"_id": "$political_preference", "count": {"$sum": 1}}}
    ])
    preferences = {}
    async for d in pref_cursor:
        preferences[d["_id"] or "unknown"] = d["count"]

    sent_cursor = db.voters.aggregate([
        {"$match": voter_filter},
        {"$group": {"_id": "$sentiment", "count": {"$sum": 1}}}
    ])
    sentiments = {}
    async for d in sent_cursor:
        sentiments[d["_id"] or "unknown"] = d["count"]

    likely = await db.voters.count_documents({**voter_filter, "likely_to_vote": True})

    return {
        "total_voters": total_voters,
        "total_booths": total_booths,
        "total_workers": total_workers,
        "total_target": total_target,
        "completion_rate": round((total_voters / total_target) * 100, 1) if total_target else 0,
        "likely_to_vote": likely,
        "preferences": preferences,
        "sentiments": sentiments,
    }

@api.get("/analytics/demographics")
async def analytics_demographics(user=Depends(get_current_user)):
    voter_filter = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        voter_filter["booth_id"] = {"$in": worker_booth_ids(user)}

    voters = await db.voters.find(voter_filter, {"_id": 0}).to_list(10000)
    age_groups = {"18-25": 0, "26-35": 0, "36-50": 0, "51-65": 0, "65+": 0, "unknown": 0}
    for v in voters:
        a = v.get("age")
        if not a:
            age_groups["unknown"] += 1
        elif a <= 25: age_groups["18-25"] += 1
        elif a <= 35: age_groups["26-35"] += 1
        elif a <= 50: age_groups["36-50"] += 1
        elif a <= 65: age_groups["51-65"] += 1
        else: age_groups["65+"] += 1

    gender = Counter([v.get("gender") or "unknown" for v in voters])
    religion = Counter([v.get("religion") or "unknown" for v in voters])
    caste = Counter([v.get("caste") or "unknown" for v in voters])
    occupation = Counter([v.get("occupation") or "unknown" for v in voters])

    return {
        "age_groups": age_groups,
        "gender": dict(gender),
        "religion": dict(religion),
        "caste": dict(caste),
        "occupation": dict(occupation),
        "total": len(voters),
    }

@api.get("/analytics/issues")
async def analytics_issues(user=Depends(get_current_user)):
    voter_filter = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        voter_filter["booth_id"] = {"$in": worker_booth_ids(user)}

    voters = await db.voters.find(voter_filter, {"_id": 0, "issues": 1}).to_list(10000)
    counter = Counter()
    for v in voters:
        for i in v.get("issues") or []:
            counter[i] += 1
    return {"issues": [{"issue": k, "count": v} for k, v in counter.most_common(15)]}

@api.get("/analytics/booth-stats")
async def booth_stats(user=Depends(get_current_user)):
    booth_filter = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        booth_filter["id"] = {"$in": worker_booth_ids(user)}

    booths = await db.booths.find(booth_filter, {"_id": 0}).to_list(1000)
    out = []
    for b in booths:
        surveyed = await db.voters.count_documents({"booth_id": b["id"]})
        supporters = await db.voters.count_documents({"booth_id": b["id"], "political_preference": "supporter"})
        out.append({
            "booth_id": b["id"],
            "name": b["name"],
            "booth_number": b["booth_number"],
            "ward": b.get("ward"),
            "target": b.get("target_voters", 0),
            "surveyed": surveyed,
            "supporters": supporters,
            "completion": round((surveyed / b["target_voters"] * 100), 1) if b.get("target_voters") else 0,
        })
    return out

@api.get("/analytics/engagement-trends")
async def engagement_trends(user=Depends(get_current_user), days: int = 14):
    voter_filter = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        voter_filter["booth_id"] = {"$in": worker_booth_ids(user)}

    voters = await db.voters.find(voter_filter, {"_id": 0, "surveyed_at": 1}).to_list(10000)
    bucket = defaultdict(int)
    for v in voters:
        s = v.get("surveyed_at")
        if not s: continue
        try:
            d = datetime.fromisoformat(s.replace("Z", "+00:00")).date().isoformat()
            bucket[d] += 1
        except Exception:
            pass
    today = datetime.now(timezone.utc).date()
    result = []
    for i in range(days - 1, -1, -1):
        day = (today - timedelta(days=i)).isoformat()
        result.append({"date": day, "count": bucket.get(day, 0)})
    return result


# ---------- SETTINGS (Party Customization) ----------
DEFAULT_SETTINGS = {
    "id": "main",
    "party_name": "VIQSO Digital Media",
    "party_short_name": "VIQSO",
    "tagline": "Connect · Create · Grow",
    "logo_url": "https://customer-assets.emergentagent.com/job_voter-hub-8/artifacts/rg7ud3ts_0BA4CCC4-7B01-4184-9028-7E2B82C624DC.png",
    "primary_color": "#8B5CF6",
    "secondary_color": "#1E90FF",
    "accent_color": "#EC4899",
    "highlight_color": "#F97316",
    "contact_email": "",
    "contact_phone": "",
    "campaign_slogan": "Win every booth",
}

class SettingsUpdate(BaseModel):
    party_name: Optional[str] = None
    party_short_name: Optional[str] = None
    tagline: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    accent_color: Optional[str] = None
    highlight_color: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    campaign_slogan: Optional[str] = None

@api.get("/settings")
async def get_settings(request: Request):
    # Try to identify org from access_key header (login page) or token (authenticated)
    org_id = DEFAULT_ORG_ID
    access_key = request.headers.get("X-Access-Key", "")
    if access_key:
        org = await db.organizations.find_one({"access_key": access_key, "active": True})
        if org:
            org_id = org["id"]
    else:
        token = request.cookies.get("access_token")
        if not token:
            h = request.headers.get("Authorization", "")
            if h.startswith("Bearer "):
                token = h[7:]
        if token:
            try:
                payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
                org_id = payload.get("org_id") or DEFAULT_ORG_ID
            except Exception:
                pass

    s = await db.settings.find_one({"org_id": org_id}, {"_id": 0})
    if not s:
        s = dict(DEFAULT_SETTINGS)
        s["org_id"] = org_id
        await db.settings.insert_one(dict(s))
    return s

@api.put("/settings")
async def update_settings(body: SettingsUpdate, admin=Depends(require_roles("admin"))):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No updates")
    updates["updated_at"] = now_iso()
    await db.settings.update_one({"org_id": admin["org_id"]}, {"$set": updates}, upsert=True)
    s = await db.settings.find_one({"org_id": admin["org_id"]}, {"_id": 0})
    return s


# ---------- BULK VOTER UPLOAD ----------
EXCEL_COLUMNS = [
    "name", "voter_id_number", "age", "gender", "address", "phone", "email",
    "caste", "religion", "occupation", "booth_number", "ward", "family_id",
    "political_preference", "sentiment", "issues", "likely_to_vote", "notes",
]

@api.get("/import/template")
async def voter_template(_user=Depends(require_roles("admin", "supervisor"))):
    """Return an Excel template for bulk voter upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Voters"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B5CF6")
    for idx, col in enumerate(EXCEL_COLUMNS, 1):
        c = ws.cell(row=1, column=idx, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(14, len(col) + 4)

    # Sample row
    sample = [
        "Amit Sharma", "VID1234567", 32, "male",
        "H.No 45, Sector 7", "+91-9876543210", "amit@example.com",
        "General", "Hindu", "Business",
        "B-101", "Ward 20 - Mumbai", "",
        "supporter", "positive",
        "Water Supply; Road Infrastructure", "true",
        "Active community member",
    ]
    for idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=idx, value=val)

    # Notes sheet
    notes = wb.create_sheet("Instructions")
    notes["A1"] = "Bulk Voter Upload — Field Guide"
    notes["A1"].font = Font(bold=True, size=14)
    instructions = [
        ("name", "Full name. Required."),
        ("voter_id_number", "Unique voter ID (EPIC). Used for smart-merge — if matches existing voter, that record will be updated."),
        ("age", "Integer."),
        ("gender", "male / female / other"),
        ("address", "Full address. Used for auto family grouping."),
        ("phone", "Contact number."),
        ("caste", "General / OBC / SC / ST / etc."),
        ("religion", "Hindu / Muslim / Christian / Sikh / etc."),
        ("occupation", "Free text."),
        ("booth_number", "Must match an existing booth (e.g., B-101). New booths will be auto-created."),
        ("ward", "Used when auto-creating booths."),
        ("family_id", "Optional. Leave blank to auto-detect via address + surname."),
        ("political_preference", "supporter / neutral / undecided / opposition"),
        ("sentiment", "positive / neutral / negative"),
        ("issues", "Semicolon-separated list, e.g. 'Water; Roads; Power'"),
        ("likely_to_vote", "true / false / yes / no"),
        ("notes", "Free text."),
    ]
    for i, (col, desc) in enumerate(instructions, 3):
        notes.cell(row=i, column=1, value=col).font = Font(bold=True)
        notes.cell(row=i, column=2, value=desc)
    notes.column_dimensions["A"].width = 22
    notes.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="voter_upload_template.xlsx"'},
    )


@api.post("/import/voters")
async def bulk_upload_voters(
    file: UploadFile = File(...),
    user=Depends(require_roles("admin", "supervisor")),
):
    from openpyxl import load_workbook
    from io import BytesIO

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Empty file")

    header = [str(h or "").strip().lower() for h in rows[0]]
    col_idx = {c: i for i, c in enumerate(header) if c in EXCEL_COLUMNS}

    if "name" not in col_idx:
        raise HTTPException(400, "Missing required column: name")

    # Preload booths by booth_number
    booths_list = await db.booths.find({}, {"_id": 0}).to_list(2000)
    booth_by_num = {b["booth_number"]: b for b in booths_list}

    inserted = 0
    updated = 0
    skipped = 0
    created_booths = 0
    errors = []

    def get(row, key):
        i = col_idx.get(key)
        if i is None or i >= len(row):
            return None
        val = row[i]
        if val is None:
            return None
        if isinstance(val, str):
            val = val.strip()
            return val if val else None
        return val

    def to_bool(v):
        if v is None: return None
        s = str(v).strip().lower()
        if s in ("true", "yes", "1", "y"): return True
        if s in ("false", "no", "0", "n"): return False
        return None

    def to_int(v):
        if v is None or v == "": return None
        try: return int(float(v))
        except: return None

    def to_list(v):
        if not v: return []
        return [x.strip() for x in str(v).replace(",", ";").split(";") if x.strip()]

    for idx, row in enumerate(rows[1:], start=2):
        try:
            name = get(row, "name")
            if not name:
                skipped += 1
                continue

            booth_num = get(row, "booth_number")
            ward = get(row, "ward")
            booth = booth_by_num.get(booth_num) if booth_num else None

            # Auto-create booth if missing
            if booth_num and not booth:
                new_booth = {
                    "id": str(uuid.uuid4()),
                    "name": f"Booth {booth_num}",
                    "booth_number": booth_num,
                    "ward": ward or "Unassigned",
                    "constituency": "",
                    "location": "",
                    "target_voters": 0,
                    "supervisor_id": None,
                    "assigned_workers": [],
                    "org_id": user["org_id"],
                    "created_at": now_iso(),
                }
                await db.booths.insert_one(dict(new_booth))
                booth_by_num[booth_num] = new_booth
                booth = new_booth
                created_booths += 1

            if not booth:
                errors.append({"row": idx, "error": "No booth_number — row skipped"})
                skipped += 1
                continue

            # Worker can only import for their assigned booths
            if user["role"] == "worker" and booth["id"] not in worker_booth_ids(user):
                errors.append({"row": idx, "error": "Booth not in your scope"})
                skipped += 1
                continue

            surname = extract_surname(name)
            address = get(row, "address") or ""
            family_id = get(row, "family_id") or auto_family_id(address, surname)

            voter_doc = {
                "name": name,
                "voter_id_number": get(row, "voter_id_number"),
                "age": to_int(get(row, "age")),
                "gender": get(row, "gender"),
                "address": address,
                "phone": get(row, "phone"),
                "email": get(row, "email"),
                "caste": get(row, "caste"),
                "religion": get(row, "religion"),
                "occupation": get(row, "occupation"),
                "political_preference": get(row, "political_preference"),
                "sentiment": get(row, "sentiment"),
                "issues": to_list(get(row, "issues")),
                "likely_to_vote": to_bool(get(row, "likely_to_vote")),
                "notes": get(row, "notes"),
                "booth_id": booth["id"],
                "surname": surname,
                "family_id": family_id,
                "org_id": user["org_id"],
                "custom_fields": {},
            }

            # Smart merge by voter_id_number (within same org)
            vid = voter_doc.get("voter_id_number")
            existing = await db.voters.find_one({"voter_id_number": vid, "org_id": user["org_id"]}) if vid else None
            if existing:
                voter_doc["updated_at"] = now_iso()
                await db.voters.update_one({"id": existing["id"]}, {"$set": voter_doc})
                updated += 1
            else:
                voter_doc["id"] = str(uuid.uuid4())
                voter_doc["surveyed_by"] = user["id"]
                voter_doc["surveyed_by_name"] = user["name"]
                voter_doc["surveyed_at"] = now_iso()
                await db.voters.insert_one(voter_doc)
                inserted += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
            skipped += 1

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "created_booths": created_booths,
        "total_rows": len(rows) - 1,
        "errors": errors[:20],
    }


# ---------- SEGREGATION / GROUPING ----------
def voter_scope_filter(user: dict) -> Dict[str, Any]:
    """Return filter dict to scope voters by user role + organization."""
    base = {"org_id": user["org_id"]}
    if user["role"] == "worker":
        base["booth_id"] = {"$in": worker_booth_ids(user)}
    return base


@api.get("/segregation/{group_by}")
async def segregate_voters(
    group_by: str,
    user=Depends(get_current_user),
    booth_id: Optional[str] = None,
    ward: Optional[str] = None,
    political_preference: Optional[str] = None,
    sentiment: Optional[str] = None,
    caste: Optional[str] = None,
    religion: Optional[str] = None,
    gender: Optional[str] = None,
):
    valid_keys = {"caste", "religion", "surname", "gender", "occupation", "political_preference", "sentiment", "age_group", "family_id", "booth_id", "ward"}
    if group_by not in valid_keys:
        raise HTTPException(400, f"Invalid group_by. Must be one of: {valid_keys}")

    base_filter = voter_scope_filter(user)
    if booth_id: base_filter["booth_id"] = booth_id
    if political_preference: base_filter["political_preference"] = political_preference
    if sentiment: base_filter["sentiment"] = sentiment
    if caste: base_filter["caste"] = caste
    if religion: base_filter["religion"] = religion
    if gender: base_filter["gender"] = gender

    # If ward filter: resolve to booth_ids first (org-scoped)
    if ward:
        booth_ids = [b["id"] async for b in db.booths.find({"ward": ward, "org_id": user["org_id"]}, {"id": 1, "_id": 0})]
        if "booth_id" in base_filter and isinstance(base_filter["booth_id"], dict) and "$in" in base_filter["booth_id"]:
            base_filter["booth_id"]["$in"] = list(set(base_filter["booth_id"]["$in"]) & set(booth_ids))
        else:
            base_filter["booth_id"] = {"$in": booth_ids}

    voters = await db.voters.find(base_filter, {"_id": 0}).to_list(20000)

    if group_by == "age_group":
        groups = defaultdict(list)
        for v in voters:
            a = v.get("age")
            if not a: key = "Unknown"
            elif a <= 25: key = "18-25"
            elif a <= 35: key = "26-35"
            elif a <= 50: key = "36-50"
            elif a <= 65: key = "51-65"
            else: key = "65+"
            groups[key].append(v)
    else:
        groups = defaultdict(list)
        for v in voters:
            key = v.get(group_by) or "Unknown"
            groups[str(key)].append(v)

    # Resolve booth_id labels (org-scoped)
    booth_label_map = {}
    if group_by in ("booth_id",):
        all_booths = await db.booths.find({"org_id": user["org_id"]}, {"_id": 0}).to_list(2000)
        booth_label_map = {b["id"]: f"{b['booth_number']} · {b['name']}" for b in all_booths}

    result = []
    for k, vs in groups.items():
        supporters = sum(1 for v in vs if v.get("political_preference") == "supporter")
        opposition = sum(1 for v in vs if v.get("political_preference") == "opposition")
        likely = sum(1 for v in vs if v.get("likely_to_vote") is True)
        label = booth_label_map.get(k, k) if booth_label_map else k
        result.append({
            "key": k,
            "label": label,
            "count": len(vs),
            "supporters": supporters,
            "opposition": opposition,
            "likely_to_vote": likely,
        })
    result.sort(key=lambda r: -r["count"])
    return {"group_by": group_by, "total": len(voters), "groups": result}


@api.get("/families")
async def list_families(
    user=Depends(get_current_user),
    booth_id: Optional[str] = None,
    ward: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 200,
):
    base_filter = voter_scope_filter(user)
    if booth_id: base_filter["booth_id"] = booth_id

    if ward:
        booth_ids = [b["id"] async for b in db.booths.find({"ward": ward, "org_id": user["org_id"]}, {"id": 1, "_id": 0})]
        if "booth_id" in base_filter and isinstance(base_filter["booth_id"], dict):
            base_filter["booth_id"]["$in"] = list(set(base_filter["booth_id"]["$in"]) & set(booth_ids))
        else:
            base_filter["booth_id"] = {"$in": booth_ids}

    voters = await db.voters.find(base_filter, {"_id": 0}).to_list(20000)
    families = defaultdict(list)
    for v in voters:
        fid = v.get("family_id") or "ungrouped"
        families[fid].append(v)

    result = []
    for fid, members in families.items():
        if fid == "ungrouped" or len(members) < 1:
            continue
        if search:
            s = search.lower()
            if not any(s in (m.get("name") or "").lower() or s in (m.get("address") or "").lower() or s in (m.get("surname") or "").lower() for m in members):
                continue
        surname = members[0].get("surname") or extract_surname(members[0].get("name", ""))
        address = members[0].get("address") or ""
        supporters = sum(1 for m in members if m.get("political_preference") == "supporter")
        result.append({
            "family_id": fid,
            "surname": surname,
            "address": address,
            "size": len(members),
            "supporters": supporters,
            "booth_id": members[0].get("booth_id"),
            "members": [
                {"id": m["id"], "name": m["name"], "age": m.get("age"), "gender": m.get("gender"),
                 "political_preference": m.get("political_preference")}
                for m in members
            ],
        })
    result.sort(key=lambda f: -f["size"])
    return {"total_families": len(result), "families": result[:limit]}


# ---------- ROOT ----------
@api.get("/")
async def root():
    return {"message": "Voter CRM API"}


# ---------- ORGANIZATIONS (Super Admin only) ----------
class OrgCreate(BaseModel):
    name: str
    party_name: str
    access_key: Optional[str] = None
    admin_email: str
    admin_password: str
    admin_name: str = "Administrator"

@api.get("/orgs")
async def list_orgs(_=Depends(require_super_admin)):
    orgs = await db.organizations.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    for o in orgs:
        o["user_count"] = await db.users.count_documents({"org_id": o["id"]})
        o["voter_count"] = await db.voters.count_documents({"org_id": o["id"]})
    return orgs

@api.post("/orgs")
async def create_org(body: OrgCreate, _=Depends(require_super_admin)):
    import secrets
    access_key = body.access_key or f"VIQSO-{secrets.token_urlsafe(8).upper().replace('_', '').replace('-', '')[:10]}"
    if await db.organizations.find_one({"access_key": access_key}):
        raise HTTPException(400, "Access key already exists")
    org_id = str(uuid.uuid4())
    org = {
        "id": org_id,
        "name": body.name,
        "party_name": body.party_name,
        "access_key": access_key,
        "active": True,
        "created_at": now_iso(),
    }
    await db.organizations.insert_one(org)

    # Create initial admin for this org
    admin_email = body.admin_email.lower().strip()
    if await db.users.find_one({"email": admin_email, "org_id": org_id}):
        raise HTTPException(400, "Admin email already exists for this org")
    await db.users.insert_one({
        "id": str(uuid.uuid4()),
        "email": admin_email,
        "password_hash": hash_password(body.admin_password),
        "name": body.admin_name,
        "role": "admin",
        "phone": None,
        "booth_id": None,
        "assigned_booth_ids": [],
        "active": True,
        "org_id": org_id,
        "created_at": now_iso(),
    })

    # Default settings for this org
    settings_doc = dict(DEFAULT_SETTINGS)
    settings_doc["org_id"] = org_id
    settings_doc["party_name"] = body.party_name
    settings_doc["party_short_name"] = body.party_name.split()[0] if body.party_name else "Party"
    await db.settings.insert_one(settings_doc)

    org.pop("_id", None)
    return {
        "org": org,
        "admin": {"email": admin_email, "password": body.admin_password, "role": "admin"},
        "next_steps": "Share the access_key + admin credentials with the client. They sign in at /login with all three.",
    }

@api.patch("/orgs/{org_id}")
async def update_org(org_id: str, body: dict, _=Depends(require_super_admin)):
    allowed = {"name", "party_name", "active", "access_key"}
    updates = {k: v for k, v in body.items() if k in allowed}
    if not updates:
        raise HTTPException(400, "No valid fields to update")
    await db.organizations.update_one({"id": org_id}, {"$set": updates})
    o = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    return o

@api.delete("/orgs/{org_id}")
async def delete_org(org_id: str, _=Depends(require_super_admin)):
    if org_id == DEFAULT_ORG_ID:
        raise HTTPException(400, "Cannot delete default org")
    await db.organizations.update_one({"id": org_id}, {"$set": {"active": False}})
    return {"ok": True, "soft_deleted": True}


# ---------- SEED ----------
SAMPLE_ISSUES = ["Water Supply", "Road Infrastructure", "Healthcare", "Education", "Employment", "Sanitation", "Power Cuts", "Public Transport", "Crime & Safety", "Inflation"]
SAMPLE_CASTES = ["General", "OBC", "SC", "ST"]
SAMPLE_RELIGIONS = ["Hindu", "Muslim", "Christian", "Sikh", "Other"]
SAMPLE_OCC = ["Farmer", "Teacher", "Business", "Government Employee", "Daily Wage", "Student", "Homemaker", "Retired"]
SAMPLE_PREFS = ["supporter", "supporter", "neutral", "opposition", "undecided"]
SAMPLE_SENTS = ["positive", "positive", "neutral", "negative"]

async def seed_data():
    # Indexes
    # Drop legacy unique-email-only index (now scoped by org)
    try:
        await db.users.drop_index("email_1")
    except Exception:
        pass
    await db.users.create_index([("email", 1), ("org_id", 1)], unique=True)
    await db.booths.create_index("booth_number", unique=False)
    await db.booths.create_index("org_id")
    await db.voters.create_index("booth_id")
    await db.voters.create_index("org_id")
    await db.voters.create_index("family_id")
    await db.voters.create_index("surname")
    await db.organizations.create_index("access_key", unique=True)
    await db.login_attempts.create_index("at")
    await db.settings.create_index("org_id")

    # Ensure default organization
    if not await db.organizations.find_one({"id": DEFAULT_ORG_ID}):
        await db.organizations.insert_one({
            "id": DEFAULT_ORG_ID,
            "name": "VIQSO Demo Organization",
            "party_name": "VIQSO Digital Media",
            "access_key": DEFAULT_ACCESS_KEY,
            "active": True,
            "created_at": now_iso(),
        })

    # Backfill org_id for ALL existing docs that lack it (one-time migration)
    for coll_name in ("users", "booths", "voters", "visits"):
        coll = db[coll_name]
        await coll.update_many({"org_id": {"$exists": False}}, {"$set": {"org_id": DEFAULT_ORG_ID}})

    # Migrate legacy settings doc (id="main") to org-scoped
    legacy_settings = await db.settings.find_one({"id": "main"})
    if legacy_settings and "org_id" not in legacy_settings:
        await db.settings.update_one({"id": "main"}, {"$set": {"org_id": DEFAULT_ORG_ID}})

    # Ensure default org's settings doc
    if not await db.settings.find_one({"org_id": DEFAULT_ORG_ID}):
        s = dict(DEFAULT_SETTINGS)
        s["org_id"] = DEFAULT_ORG_ID
        await db.settings.insert_one(s)

    # Backfill surname/family_id for legacy voters
    legacy = await db.voters.find({"$or": [{"surname": {"$exists": False}}, {"family_id": {"$exists": False}}]}, {"_id": 0}).to_list(20000)
    for v in legacy:
        sn = v.get("surname") or extract_surname(v.get("name", ""))
        fid = v.get("family_id") or auto_family_id(v.get("address", ""), sn)
        await db.voters.update_one({"id": v["id"]}, {"$set": {"surname": sn, "family_id": fid}})

    # Admin user
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@crm.com")
    admin_pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing_admin = await db.users.find_one({"email": admin_email, "org_id": DEFAULT_ORG_ID})
    if not existing_admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "email": admin_email,
            "password_hash": hash_password(admin_pw),
            "name": "Admin User", "role": "admin", "phone": "+91-9000000001",
            "booth_id": None, "assigned_booth_ids": [], "active": True,
            "org_id": DEFAULT_ORG_ID,
            "created_at": now_iso(),
        })
    else:
        if not verify_password(admin_pw, existing_admin["password_hash"]):
            await db.users.update_one({"email": admin_email, "org_id": DEFAULT_ORG_ID}, {"$set": {"password_hash": hash_password(admin_pw)}})

    # Check if already seeded
    if await db.booths.count_documents({}) > 0:
        logger.info("Data already seeded")
        return

    # Booths
    wards = ["Ward 12 - North Zone", "Ward 18 - East Zone", "Ward 22 - Central Zone", "Ward 27 - South Zone"]
    booths = []
    for i in range(1, 9):
        bid = str(uuid.uuid4())
        booths.append({
            "id": bid,
            "name": f"Booth {chr(64+i)} - {random.choice(['Public School', 'Community Hall', 'Govt Office', 'Library'])}",
            "booth_number": f"B-{100 + i}",
            "ward": random.choice(wards),
            "constituency": "Central Constituency",
            "location": f"Sector {i}, Main Road",
            "target_voters": random.randint(800, 1500),
            "supervisor_id": None,
            "assigned_workers": [],
            "org_id": DEFAULT_ORG_ID,
            "created_at": now_iso(),
        })
    await db.booths.insert_many([dict(b) for b in booths])

    # Supervisor
    sup_email = os.environ.get("SUPERVISOR_EMAIL", "supervisor@crm.com")
    sup_pw = os.environ.get("SUPERVISOR_PASSWORD", "super123")
    sup_id = str(uuid.uuid4())
    await db.users.insert_one({
        "id": sup_id, "email": sup_email,
        "password_hash": hash_password(sup_pw),
        "name": "Priya Sharma", "role": "supervisor", "phone": "+91-9000000002",
        "booth_id": None, "assigned_booth_ids": [b["id"] for b in booths[:4]],
        "active": True, "org_id": DEFAULT_ORG_ID, "created_at": now_iso(),
    })

    # Worker (login user)
    wkr_email = os.environ.get("WORKER_EMAIL", "worker@crm.com")
    wkr_pw = os.environ.get("WORKER_PASSWORD", "worker123")
    wkr_id = str(uuid.uuid4())
    await db.users.insert_one({
        "id": wkr_id, "email": wkr_email,
        "password_hash": hash_password(wkr_pw),
        "name": "Rajesh Kumar", "role": "worker", "phone": "+91-9000000003",
        "booth_id": booths[0]["id"], "assigned_booth_ids": [booths[0]["id"], booths[1]["id"]],
        "active": True, "org_id": DEFAULT_ORG_ID, "created_at": now_iso(),
    })

    # More workers
    worker_ids = [wkr_id]
    worker_names = ["Anjali Patel", "Vikram Singh", "Meera Nair", "Sandeep Verma", "Kavita Reddy"]
    for idx, nm in enumerate(worker_names):
        wid = str(uuid.uuid4())
        worker_ids.append(wid)
        await db.users.insert_one({
            "id": wid, "email": f"worker{idx+1}@crm.com",
            "password_hash": hash_password("worker123"),
            "name": nm, "role": "worker", "phone": f"+91-90000000{20+idx}",
            "booth_id": booths[idx % len(booths)]["id"],
            "assigned_booth_ids": [booths[idx % len(booths)]["id"]],
            "active": True, "org_id": DEFAULT_ORG_ID, "created_at": now_iso(),
        })

    # Sample voters
    first_names = ["Amit", "Sunita", "Ravi", "Pooja", "Sanjay", "Geeta", "Rakesh", "Neha", "Manoj", "Kiran", "Deepak", "Asha", "Vinod", "Rekha", "Suresh"]
    last_names = ["Sharma", "Verma", "Yadav", "Kumar", "Singh", "Gupta", "Patel", "Reddy", "Mishra", "Joshi", "Chauhan"]
    voters = []
    for i in range(120):
        booth = random.choice(booths)
        wid = random.choice(worker_ids)
        days_ago = random.randint(0, 13)
        ts = (datetime.now(timezone.utc) - timedelta(days=days_ago, hours=random.randint(0, 23))).isoformat()
        voters.append({
            "id": str(uuid.uuid4()),
            "booth_id": booth["id"],
            "name": f"{random.choice(first_names)} {random.choice(last_names)}",
            "voter_id_number": f"VID{random.randint(1000000, 9999999)}",
            "age": random.randint(19, 78),
            "gender": random.choice(["male", "male", "female", "female", "other"]),
            "address": f"H.No {random.randint(1, 200)}, Sector {random.randint(1, 20)}",
            "phone": f"+91-9{random.randint(100000000, 999999999)}",
            "email": None,
            "caste": random.choice(SAMPLE_CASTES),
            "religion": random.choice(SAMPLE_RELIGIONS),
            "occupation": random.choice(SAMPLE_OCC),
            "political_preference": random.choice(SAMPLE_PREFS),
            "sentiment": random.choice(SAMPLE_SENTS),
            "issues": random.sample(SAMPLE_ISSUES, k=random.randint(1, 4)),
            "likely_to_vote": random.choice([True, True, True, False, None]),
            "notes": random.choice(["Active community member", "Wants better roads", "First-time voter", "", ""]),
            "custom_fields": {},
            "org_id": DEFAULT_ORG_ID,
            "surveyed_by": wid,
            "surveyed_by_name": "Field Worker",
            "surveyed_at": ts,
        })
    if voters:
        await db.voters.insert_many(voters)

    # Visits
    visits = []
    for i in range(15):
        booth = random.choice(booths)
        wid = random.choice(worker_ids)
        day_offset = random.randint(-3, 7)
        date = (datetime.now(timezone.utc) + timedelta(days=day_offset)).isoformat()
        status = "completed" if day_offset < 0 else "scheduled"
        visits.append({
            "id": str(uuid.uuid4()),
            "booth_id": booth["id"],
            "worker_id": wid,
            "scheduled_date": date,
            "status": status,
            "voters_contacted": random.randint(0, 50) if status == "completed" else 0,
            "notes": random.choice(["Morning visit", "Evening canvassing", "Door-to-door survey", ""]),
            "org_id": DEFAULT_ORG_ID,
            "created_at": now_iso(),
        })
    if visits:
        await db.visits.insert_many(visits)

    logger.info(f"Seeded: {len(booths)} booths, {len(voters)} voters, {len(visits)} visits, {len(worker_ids)+2} users")


@app.on_event("startup")
async def startup():
    await seed_data()

@app.on_event("shutdown")
async def shutdown():
    client.close()


# Include router
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_origin_regex=".*",
    allow_methods=["*"],
    allow_headers=["*"],
)
